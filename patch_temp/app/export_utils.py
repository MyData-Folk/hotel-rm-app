from io import BytesIO
from typing import List, Dict, Any, Tuple
from datetime import datetime

def _normalize_tabular(rows: Any, columns: List[str] | None = None) -> Tuple[List[str], List[List[Any]]]:
    if not rows:
        return (columns or [], [])
    if isinstance(rows[0], dict):
        keys = columns or list(rows[0].keys())
        data = [[r.get(k) for k in keys] for r in rows]
        return (keys, data)
    if isinstance(rows[0], (list, tuple)):
        if columns is None:
            max_len = max(len(r) for r in rows)
            columns = [f"col_{i+1}" for i in range(max_len)]
        return (columns, [list(r) for r in rows])
    return (columns or ["value"], [[str(x)] for x in rows])

def build_excel_bytes(kind: str, hotel_id: str, rows: Any, columns: List[str] | None = None, metadata: Dict[str, Any] | None = None) -> bytes:
    import pandas as pd
    from openpyxl.utils import get_column_letter
    headers, data = _normalize_tabular(rows, columns)
    df = pd.DataFrame(data, columns=headers or None)
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        sheet = (kind or "data").capitalize()
        df.to_excel(writer, index=False, sheet_name=sheet)
        wb = writer.book
        ws = writer.sheets[sheet]
        for i, col in enumerate(df.columns, start=1):
            max_len = max([len(str(col))] + [len(str(x)) for x in df[col].astype(str).values]) if not df.empty else len(str(col))
            ws.column_dimensions[get_column_letter(i)].width = min(max(10, max_len + 2), 60)
        meta = metadata or {}
        info_ws = wb.create_sheet("Meta")
        info_ws.append(["hotel_id", hotel_id])
        info_ws.append(["kind", kind])
        info_ws.append(["generated_at", datetime.utcnow().isoformat() + "Z"])
        for k, v in meta.items():
            info_ws.append([k, str(v)])
    return bio.getvalue()

def build_pdf_bytes(kind: str, hotel_id: str, rows: Any, columns: List[str] | None = None, metadata: Dict[str, Any] | None = None) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    headers, data = _normalize_tabular(rows, columns)
    bio = BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=landscape(A4), title=f"{kind} - {hotel_id}")
    styles = getSampleStyleSheet()
    story = []
    title = f"{kind.capitalize()} — Hôtel: {hotel_id}"
    story.append(Paragraph(title, styles["Title"]))
    story.append(Spacer(1, 12))
    meta = metadata or {}
    meta_text = ", ".join([f"{k}: {v}" for k, v in meta.items()]) if meta else "—"
    story.append(Paragraph(f"<b>Méta:</b> {meta_text}", styles["Normal"]))
    story.append(Spacer(1, 12))
    table_data = [headers] + data if headers else data
    table = Table(table_data, repeatRows=1)
    table_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#e5e7eb')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#111827')),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('GRID', (0,0), (-1,-1), 0.25, colors.HexColor('#d1d5db')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f9fafb')]),
    ])
    table.setStyle(table_style)
    story.append(table)
    doc.build(story)
    return bio.getvalue()
